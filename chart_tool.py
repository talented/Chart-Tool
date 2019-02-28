
import os
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from collections import Counter, defaultdict

import numpy as np
import warnings

import re
import operator

import matplotlib
matplotlib.use("Agg")
#matplotlib.use("WXAgg")
import matplotlib.pyplot as plt
#from matplotlib import font_manager as fm
#from mpl_toolkits.mplot3d import Axes3D
#from pprint import pprint
#import pylab as pl
#from pylab import figure, axes, pie, title, show
#import matplotlib.image as mpimg
#import scipy.stats as stats
#import matplotlib.path as mpath
#import matplotlib.lines as mlines
#import matplotlib.patches as mpatches
#from matplotlib.collections import PatchCollection
#import Tkinter

#-------------------------------------------------------------------------------------#

#type(wb)
# reference expressions to be checked when necessary DON'T DELETE!!!
#cell_swb0_a1 = swb0.cell('A1')
#cell_swb0_a1.value = 'Wrote to cell in 1st sheet.'
#writer = ExcelWriter(workbook=swb)
#writer.save('example.xlsx')
#print sheet['A1'].column # A
#print sheet['A1'].row # 1
#print sheet.cell(row=1, column=2).value
#for one in range(97,123):
#print chr(one).upper()
#print act_sheet.max_row
#swb0.cell('A' + str(j)).value
#key = act_sheet['A'+ str(row)].value

#-------------------------------------------------------------------------------------#    

# change this!
DEFAULT = 'C:\Users\username\.spyder2'


def current_dir(cwd):
    path = cwd + '/charts'
    if not os.path.exists(path):
       os.makedirs(path)
       print 'Directory is created'
    os.chdir(path)
        
def read_excelfile(readfilename): 

    warnings.simplefilter("ignore")
    wb = openpyxl.load_workbook(readfilename)
    sheet = wb.get_active_sheet()
    return sheet
    
# get a dictionary on format below
# {'user_name': {'department': 'dep', 'quantity': number}} 
# example {'MATHONET Meike': {'Originator Department': 'POAE2', quantity: '34'} }  
def get_user_dict(readfilename, columnname, filter_header):
    act_sheet = read_excelfile(readfilename)
    table = []
    item_dict = {}
#    create a list of uppercase letters for A-Z
    col_header = [chr(one).upper() for one in range(97,123)]
    
    for idx, header in enumerate(col_header):
        head = header + str(1)    
        if act_sheet[head].value == columnname:
            for j in range(2,act_sheet.max_row+1):
                items = act_sheet[header + str(j)].value.upper()
                departments = act_sheet[col_header[idx+1] + str(j)].value
                table.append(items)
                item_dict.setdefault(items, {})
                item_dict[items].setdefault(filter_header, departments)
                
    
    amounts = Counter(table) 
    for k,v in amounts.items():
        item_dict[k]['Quantity'] = v
   
    return item_dict
    
#-------------------------------------------------------------------------------------#

def get_column_dict(readfilename, columnname):

    act_sheet = read_excelfile(readfilename)
    headers = {c.value: idx+1  for idx, c in enumerate(list(act_sheet.rows)[0])}
    col = headers[columnname] - 1
    table = []
    for row in act_sheet.iter_rows():
        table.append(unicode(row[col].value))
    table.pop(0)
    return Counter(table)

#-------------------------------------------------------------------------------------#
    
def change_dir():
    cwd = os.getcwd()
    path = cwd + '\\charts'
    if not os.path.exists(path):
        os.makedirs(path)
#        print 'Directory is created'  
    os.chdir(path)

#-------------------------------------------------------------------------------------#
    
def write_to(writefilename, dictname, headers):
    swb = Workbook()
    swb0 = swb.worksheets[0]
    swb.title = str(dictname)
    swb0.append(headers)
    table = [dictname.keys(), dictname.values()]
    
    for i in range(len(table)):
        for idx, row in enumerate(table[i]):
            swb0.cell(column=i+1, row=idx+2).value = row
            
    writer = ExcelWriter(workbook=swb)
    writer.save(writefilename + '.xlsx')
    print 'filename has been created'

#-------------------------------------------------------------------------------------#
     
def write_to_excel(writefilename, dictname, headers):
    swb = Workbook()
    swb0 = swb.worksheets[0]
    swb0.title = 'Sheet1'

    swb0.append(headers)
    keys = list(dictname.keys())
    for idx, val in enumerate(keys):    
        swb0.cell(column=1, row=idx+2).value = val
     
    values = dictname.values()
    for idx, valDict in enumerate(values):
        swb0.cell(column=2, row=idx+2).value = valDict[headers[1]]
        swb0.cell(column=3, row=idx+2).value = valDict[headers[2]]
     
    change_dir()
    writer = ExcelWriter(workbook=swb)
    writer.save(writefilename + '.xlsx')
    print 'filename has been created'

#-------------------------------------------------------------------------------------#

def write_to_excel_deps(writefilename, dictname, headers):
    swb = Workbook()
    swb0 = swb.worksheets[0]
    swb0.title = 'Sheet1'

    swb0.append(headers)

    for idx, valDict in enumerate(collapse_dict(dictname)):
        swb0.cell(column=1, row=idx+2).value = valDict[headers[0]]
        swb0.cell(column=2, row=idx+2).value = valDict[headers[1]]
     
    change_dir()
    writer = ExcelWriter(workbook=swb)
    writer.save(writefilename + '.xlsx')
    print 'filename has been created'
    
#-------------------------------------------------------------------------------------#   
    
def add_sheet_excel(writefilename, sheet, dictname, headers):
    pass

#-------------------------------------------------------------------------------------#                             

def collapse_dict(dictionary):
    user = dictionary.values()
    newDict = {}
    
    for item in user:
        n, q = item.values()
        newDict[n] = newDict.get(n, 0) + q   

    user = [{'Originator Department':k, 'Quantity':v} for k,v in newDict.items()]
    return user

#-------------------------------------------------------------------------------------#

# get a dictionary on format below with procent values
#[{'Originator Department': u'PQAT1', 'Quantity': 3, 'procent': '0.8'},
# {'Originator Department': u'POAC11', 'Quantity': 35, 'procent': '9.3'},
# {'Originator Department': u'POAC41', 'Quantity': 16, 'procent': '4.3'},..]
def get_procent(readfilename):
    orig_concession = get_user_dict(readfilename, 'Originator Name', 'Originator Department')
    newList = collapse_dict(orig_concession)
    count = 0
    for item in newList:
        count += item['Quantity']
                
    for item in newList:
        q = item['Quantity']
        procent = q * 100 / float(count)
        item['procent'] = format(procent, '0.1f')
    
    return newList

#-------------------------------------------------------------------------------------#

def get_rto_reasons(readfilename, columnname, filter_header=None):
    act_sheet = read_excelfile(readfilename)

    headers = {c.value: idx+1  for idx, c in enumerate(list(act_sheet.rows)[0])}

    col = headers[columnname] - 1
    
    if filter_header != None:
        dep_col = headers[filter_header] - 1
        reasons = defaultdict(list)
        for row in act_sheet.iter_rows():
            reason = row[col].value
            dep = row[dep_col].value
            reasons[reason].append(dep)
        return reasons    
    else:
        table = []
        for row in act_sheet.iter_rows():
            table.append(unicode(row[col].value))
        table.pop(0)
        return table

#-------------------------------------------------------------------------------------#
        
def get_plot_style():
    
#    STYLES
    #    print plt.style.available
    plt.style.use('ggplot')
#    plt.rcParams.update(plt.rcParamsDefault)
#    print plt.style.available
#    with plt.style.context(('ggplot')):

#    PARAMS
   # get all available rcParams
#    pprint(plt.rcParams.keys())
    plt.rcParams['grid.alpha'] = 0.2
    plt.rcParams['grid.color'] = 'gray'
    plt.rcParams['axes.facecolor'] = 'white'
    plt.rcParams['axes.edgecolor'] = 'white'
    plt.rcParams['text.antialiased'] = True
    plt.rcParams['axes.grid'] = True
    plt.rcParams['polaraxes.grid'] = False
    plt.rcParams['axes.formatter.useoffset'] = False
    plt.rcParams['contour.negative_linestyle'] = 'solid'
    plt.tick_params(top='off', bottom='off', left='off', right='off')
    
    fig_size = plt.rcParams['figure.figsize']
    fig_size[0] = 12
    fig_size[1] = 6
    plt.rcParams['figure.figsize'] = fig_size
    
#    FONTS
    font = {'family': 'sans-serif',
        'weight' : 'normal',
        }
#
    plt.rc('font', **font)

#    SMALL_SIZE = 8
    MEDIUM_SIZE = 16
    BIGGER_SIZE = 16
    TITLE_SIZE = 26
    
#    plt.rc('font', size=BIGGER_SIZE)          # controls default text sizes
    plt.rc('axes', titlesize=TITLE_SIZE)     # fontsize of the axes title
    plt.rc('axes', labelsize=BIGGER_SIZE)    # fontsize of the x and y labels
    plt.rc('xtick', labelsize=MEDIUM_SIZE)    # fontsize of the tick labels
    plt.rc('ytick', labelsize=MEDIUM_SIZE)    # fontsize of the tick labels
#    plt.rc('legend', fontsize=TITLE_SIZE)    # legend fontsize
#    plt.rc('xtick', labelsize=14) 
#    plt.rc('ytick', labelsize=14) 
    
#    ADJUSTMENTS
    plt.subplots_adjust(left=1, bottom=None, right=2, top=None, wspace=2, hspace=None)
    
#    LEGEND   
    plt.rcParams['legend.fancybox'] = True
    plt.rcParams['legend.shadow'] = True
    plt.rcParams['legend.borderpad'] = 1.5
    plt.rcParams['legend.handlelength'] = 2

#-------------------------------------------------------------------------------------#   

def draw_reason(filename, yaxis, xaxis, field, currdir=DEFAULT, titleS=True, titleT=None, angle=45, percentage=True,
                xlegend=None, legendloc='Left', Bsorted=True, xlabel=True, xlabelT=None, ylabel=True, ylabelT=None,
                Bbcolor='Dark Blue', Bpcolor='Indian Red', hidden=None):
    os.chdir(currdir)
    reasons_dict = get_rto_reasons(filename, yaxis, xaxis)
    counted = Counter(reasons_dict[field])
    sorted_counted = sorted(counted.items(), key=operator.itemgetter(1))

    if Bsorted:
        xitems = [v[0] for v in sorted_counted]
        yitems = [v[1] for v in sorted_counted]
    else:
        xitems = counted.keys()
        yitems = counted.values()
#
    # calculate percentage
    percentages = []
    total = sum(yitems)
    for i in yitems:
        procent = i * 100 / float(total)
        percentages.append(procent)
        
    # DRAW
    fig, ax = plt.subplots()
    get_plot_style()
   
    y_pos = np.arange(len(xitems)) + 0.5
    offset = 3
    width = 1
    if percentage:
        ax.set_xmargin(0.03)
    else:
        ax.set_xmargin(0.01)
        
    # color selection
    #-----------------------#
    color_list = ['Dark Blue', 'Blue', 'Shutter Blue', 'Dark Green', 'Yellow', 'Pink']
    hex_list = ['#262673','#137ed9', '#0c9fc6', '#405d27', '#feb236', '#d64161']

    for idx, c in enumerate(color_list):
        if Bbcolor == c:
            barcolor = hex_list[idx]
        elif Bbcolor == 'Color Mix':
            barcolor = hex_list
            
    color_listP = ['Indian Red', 'Red', 'Orange', 'Purple', 'Light Green', 'Light Blue']
    hex_listP = ['indianred', '#c94c4c', '#ff7b25', '#6b5b95', '#b5e7a0', '#80ced6']
    for idx, c in enumerate(color_listP):
        if Bpcolor == c:
            percolor = hex_listP[idx]
    
    #plt.rcParams['axes.linewidth'] = 1
    [i.set_linewidth(0.1) for i in ax.spines.itervalues()]
    
    #---------------------------#
  
    barNum = ax.bar(y_pos * offset, yitems, width, color=barcolor,
            alpha=1, linewidth=1, edgecolor='#131339')                                             
    
     # Label settings
    if xlabel:
        if xlabelT != None:        
            ax.set_xlabel(xlabelT, color='indianred', labelpad=20)
        else:
            ax.set_xlabel(xaxis, color='indianred', labelpad=20)
            
    if ylabel:
        if ylabelT != None:        
            ax.set_ylabel(ylabelT, color='indianred', labelpad=10)
        else:
            ax.set_ylabel('Quantity', color='indianred', labelpad=10)
    
    # Title settings
    if titleS:
        if titleT != None:
            ax.set_title(titleT, y=1.08)
        else:
            q = str(sum([int(q) for q in yitems]))
            titleWithQ = field + ' (total: ' + q + ')'
            ax.set_title(titleWithQ, y=1.08)
#    ax.set_ylim([ymin, ymax])
            
    # bar settings
    if angle == 90:
        ax.set_xticks(y_pos * offset + 0.5)
        ax.set_xticklabels(xitems, 2, rotation=angle)
    else:
        ax.set_xticks(y_pos * offset)
        ax.set_xticklabels(xitems, rotation=angle, ha='right', rotation_mode = 'anchor')
  
    #for lab in ax.get_xticklabels():
     #   lab.set_rotation(angle)

    fontD = {'family': 'sans-serif', 'weight': 'bold', 'size': 14}
    labelsize = 16
    if len(xitems) > 20:
        labelsize = 14
        fontD['size'] = 12
    if len(xitems) > 30:
        labelsize = 12
        fontD['size'] = 10
    if len(xitems) > 40:
        labelsize = 10
        fontD['size'] = 8
    if len(xitems) > 60:
        labelsize = 6
        fontD['size'] = 6
    if len(xitems) > 90:
        labelsize = 4
        fontD['size'] = 4
        
    for ticks in ax.get_xticklabels():
        ticks.set_fontsize(labelsize)
          
     # text over the bar
    for bar in barNum:
        height = bar.get_height()
        ax.text(bar.get_x()+bar.get_width()/2., 1.01*height,
                             '%d' % int(height), ha='center', va='bottom', fontdict=fontD)
          
    if percentage:    
        ax.set_xticks(y_pos * offset + 1.1)
        
        barPro = ax.bar(y_pos * offset + 1.2, percentages, width, color=percolor,
                         alpha=1, linewidth=1, edgecolor='#3d1b1b')
                         
        if hidden == 'ozzy':
            fontP = {'family': 'fantasy', 'weight': 'bold', 'size': 10}
            if len(xitems) > 10:
                fontP['size'] = 8
            if len(xitems) > 20:
                fontP['size'] = 6
            if len(xitems) > 30:
                fontP['size'] = 4
            if len(xitems) > 40:
                fontP['size'] = 4
            if len(xitems) > 60:
                fontP['size'] = 2
            if len(xitems) > 90:
                fontP['size'] = 2
    #        text over the bar
            for bar in barPro:
                height = bar.get_height()
                ax.text(bar.get_x()+bar.get_width()/2., 1.01*height,
                                     '%.1f' % float(height), ha='center', va='bottom', fontdict=fontP)
                         
         # legend settings
        if legendloc == 'Bottom':
            legendcoor = (0.5, -0.25)
            ncol = 2
            loc = 'center'
        elif legendloc == 'Left':
            legendcoor = (0.03, 1)
            ncol = 1
            loc = 'upper left'
        elif legendloc == 'Right':
            legendcoor = (1, 1)
            ncol = 1
            loc = 'upper left'
            
        legend = ax.legend((barNum[0], barPro[0]), (xlegend + ' Quantity', xlegend + ' Percentage'),
                           loc=loc, bbox_to_anchor=legendcoor, ncol=ncol)
        legend.get_frame().set_edgecolor('#131339')
        legend.get_frame().set_alpha(0.9)
    
#    savename
    regex = re.compile('[^a-zA-Z0-9]')
    altered = regex.sub('_', field)
    savename = altered + '_Bar.png'
    
#    plt.show()
    current_dir(currdir)
    fig.savefig(savename, dpi=199, bbox_inches='tight')
    plt.close(fig)
 
#-------------------------------------------------------------------------------------#
           
def draw_xy(readfilename, yaxis, titleS=True, titleT=None, currdir=DEFAULT, angle=90, percentage=True,
            legendloc='Left', Bsorted=True, xlabel=True, xlabelT=None, ylabel=True, ylabelT=None, xlegend=None,
            Bbcolor='Dark Blue', Bpcolor='Indian Red', hidden=None):

    os.chdir(currdir)
    drawDict = get_column_dict(readfilename, yaxis)
    
    sorted_drawDict = sorted(drawDict.items(), key=operator.itemgetter(1))
    if Bsorted:
        xitems = [v[0] for v in sorted_drawDict]
        yitems = [v[1] for v in sorted_drawDict]
    else:
        xitems = [str(x) for x in drawDict.keys()]
        yitems = drawDict.values()
        
    # calculate percentage
    percentages = []
    total = sum(yitems)
    for i in yitems:
        procent = i * 100 / float(total)
        percentages.append(procent)

    # DRAW
    fig, ax = plt.subplots()
    get_plot_style()
    
    y_pos = np.arange(len(xitems)) + 0.5
    offset = 3
    width = 1
    if percentage:
        ax.set_xmargin(0.03)
    else:
        ax.set_xmargin(0.01)
        
    # color selection
    #-----------------------#
    color_list = ['Dark Blue', 'Blue', 'Shutter Blue', 'Dark Green', 'Yellow', 'Pink']
    hex_list = ['#262673','#137ed9', '#0c9fc6', '#405d27', '#feb236', '#d64161']

    for idx, c in enumerate(color_list):
        if Bbcolor == c:
            barcolor = hex_list[idx]
        elif Bbcolor == 'Color Mix':
            barcolor = hex_list
            
    color_listP = ['Indian Red', 'Red', 'Orange', 'Purple', 'Light Green', 'Light Blue']
    hex_listP = ['indianred', '#c94c4c', '#ff7b25', '#6b5b95', '#b5e7a0', '#80ced6']
    for idx, c in enumerate(color_listP):
        if Bpcolor == c:
            percolor = hex_listP[idx]
    
    #plt.rcParams['axes.linewidth'] = 1
    [i.set_linewidth(0.1) for i in ax.spines.itervalues()]
    
    #---------------------------#

    
    barNum = ax.bar(y_pos * offset, yitems, width, color=barcolor,
            alpha=1, linewidth=1, edgecolor='#131339')

      # Label settings
    if xlabel:
        if xlabelT != None:        
            ax.set_xlabel(xlabelT, color='indianred', labelpad=20)
        else:
            ax.set_xlabel(yaxis, color='indianred', labelpad=20)
            
    if ylabel:
        if ylabelT != None:        
            ax.set_ylabel(ylabelT, color='indianred', labelpad=10)
        else:
            ax.set_ylabel('Quantity', color='indianred', labelpad=10)
    
    # Title settings
    if titleS:
        if titleT != None:
            ax.set_title(titleT, y=1.08)
        else:
            q = str(sum([int(q) for q in yitems]))
            titleWithQ = yaxis + ' (total: ' + q + ')'
            ax.set_title(titleWithQ, y=1.08)
             
    # bar settings
    if angle == 90:
        ax.set_xticks(y_pos * offset + 0.5)
        ax.set_xticklabels(xitems, 2, rotation=angle)
    else:
        ax.set_xticks(y_pos * offset)
        ax.set_xticklabels(xitems, rotation=angle, ha='right', rotation_mode = 'anchor')
  
    #for lab in ax.get_xticklabels():
     #   lab.set_rotation(angle)

    fontD = {'family': 'sans-serif', 'weight': 'bold', 'size': 14}
    labelsize = 16
    if len(xitems) > 20:
        labelsize = 14
        fontD['size'] = 12
    if len(xitems) > 30:
        labelsize = 12
        fontD['size'] = 10
    if len(xitems) > 40:
        labelsize = 10
        fontD['size'] = 8
    if len(xitems) > 60:
        labelsize = 6
        fontD['size'] = 6
    if len(xitems) > 90:
        labelsize = 4
        fontD['size'] = 4
      
    print ax.get_xticklabels()
    for ticks in ax.get_xticklabels():
        ticks.set_fontsize(labelsize)
 
    # text over the bar
    for bar in barNum:
        height = bar.get_height()
        ax.text(bar.get_x()+bar.get_width()/2., 1.01*height,
                             '%d' % int(height), ha='center', va='bottom', fontdict=fontD)                           
    
    if percentage:       
        ax.set_xticks(y_pos * offset + 1.1)
       
        barPro = ax.bar(y_pos * offset + 1.1, percentages, width, color=percolor, alpha=1,
                        linewidth=1, edgecolor='#3d1b1b')

        if hidden == 'ozzy':
            fontP = {'family': 'fantasy', 'weight': 'bold', 'size': 10}
            if len(xitems) > 10:
                fontP['size'] = 8
            if len(xitems) > 20:
                fontP['size'] = 6
            if len(xitems) > 30:
                fontP['size'] = 4
            if len(xitems) > 40:
                fontP['size'] = 4
            if len(xitems) > 60:
                fontP['size'] = 2
            if len(xitems) > 90:
                fontP['size'] = 2
    #        text over the bar
            for bar in barPro:
                height = bar.get_height()
                ax.text(bar.get_x()+bar.get_width()/2., 1.01*height,
                                     '%.1f' % float(height), ha='center', va='bottom', fontdict=fontP)
               
        # legend settings
        if legendloc == 'Bottom':
            legendcoor = (0.5, -0.25)
            ncol = 2
            loc = 'center'
        elif legendloc == 'Left':
            legendcoor = (0.03, 1)
            ncol = 1
            loc = 'upper left'
        elif legendloc == 'Right':
            legendcoor = (1, 1)
            ncol = 1
            loc = 'upper left'
            
        legend = ax.legend((barNum[0], barPro[0]), (xlegend + ' Quantity', xlegend + ' Percentage'),
                           loc=loc, bbox_to_anchor=legendcoor, ncol=ncol)
        legend.get_frame().set_edgecolor('#131339')
        legend.get_frame().set_alpha(0.9)
    
#    savename
    regex = re.compile('[^a-zA-Z0-9]')
    altered = regex.sub('_', yaxis)
    savename = altered + '_Bar.png'

#    plt.show()
    current_dir(currdir)
    fig.savefig(savename, dpi=199, bbox_inches='tight')
    plt.close(fig)

#-------------------------------------------------------------------------------------#

def get_percent(filename, columnname, divergences=None, variable=None):
    
    reasons_dict = get_rto_reasons(filename, columnname, divergences)
    
    counted = {}
    if divergences != None and variable != None:
        counted = Counter(reasons_dict[variable])
    else:
        counted = Counter(reasons_dict)
    
    xitems = counted.keys()
    yitems = counted.values()
    final = []
    
    # calculate percentage
    percentages = []
    total = sum(yitems)
    for i in yitems:
        procent = i * 100 / float(total)
        percentages.append(format(procent, '0.1f'))
       
    for idx, item in enumerate(xitems):
        new = {}
        new.setdefault('item', xitems[idx])
        new.setdefault('Quantity', yitems[idx])
        new.setdefault('Percentage', percentages[idx])
        final.append(new)
 
    sortedList = sorted(final, key=lambda k:k['Quantity'])
    return sortedList

#-------------------------------------------------------------------------------------# 

def draw_pie(filename, columnname, divergences=None, field=None, currdir=DEFAULT, pieTitle=True,
             givenTitle=None, randomColors=False, collapsed=True, exploded=False, legendloc='Bottom', procent=4,
             pctdistance=0.6, Pcolor='Blue-Green'):
   
    os.chdir(currdir)
    sortedList = get_percent(filename, columnname, divergences, field)
    percentages = [float(item['Percentage']) for item in sortedList]
    shortened = []
    finalList = []
    labels = []
    labelsQ = []
    count = 0
    times = 0
    quantities = []
    colors = ''
    
    for item in percentages:
        if item > procent:
            shortened.append(item)
        else:
            count += item
            times += 1
    
    if count > 0:
        finalList = [format(count, '0.1f')] + shortened
        deps = [item['item'] for item in sortedList][-len(finalList)+1:]
        labels = ['Others'] + deps
        qua = [item['Quantity'] for item in sortedList]
        quaint = [sum(qua[:times])] + qua[times:]
        quantities = [str(x) for x in quaint]
        
    else:
        finalList = shortened
        deps = [item['item'] for item in sortedList][-len(finalList):]
        labels = deps  
        quantities = [str(item['Quantity']) for item in sortedList]
    
    labelsQ = [l + '(' + quantities[idx] + ')' for idx, l in enumerate(labels)]
    
#    labelsS = []
#    for label in labelsQ:
#        
#        q = len(re.findall(regex, label))
#        a = label.split(' ')
#        b = a[:]
#        b.insert(2, '\n')
#        labelsS.append(' '.join(b))

    # COLOR SETTINGS
    if Pcolor == 'Blue-Green':
        colors = ['#191970', '#001CF0','#0055D4',
    '#0071C6','#008DB8','#00AAAA','#00C69C','#00E28E','#00FF80']
    elif Pcolor == 'Dark':
        colors = ['#267373', '#265973', '#264073', '#262673', '#402673', '#592673', '#732673',
           '#732659', '#732640', '#732626', '#734026']
    elif Pcolor == 'Color Mix':       
#    colors = random.sample(colorlst, len(shortened)) 
        colors=plt.style.library['bmh']['axes.color_cycle']
#         RANDOM COLORS
#        N = 100
#        HSV_tuples = [(x*1.0/N, 0.5, 0.5) for x in xrange(N)]
#        hex_out = []
#        for rgb in HSV_tuples:
#            rgb = map(lambda x: int(x*255),colorsys.hsv_to_rgb(*rgb))
#            hex_out.append("".join(map(lambda x: chr(x).encode('hex'),rgb)))
#        colors = ['#' + item for item in hex_out]
#        colors = [hex for name, hex in matplotlib.colors.cnames.iteritems()]

    csfont = {'fontname':'helvetica'}
    
    # Make a square figure and axes
    fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(aspect="equal"))
    plt.rcParams['axes.facecolor'] = 'white'
    plt.rcParams['text.color'] = 'black'
    plt.rcParams['axes.labelsize'] = 30
#    pprint(plt.rcParams.keys())
    if exploded:
        explode = ((0,) * (len(finalList)-1)) + (0.1,) 
    else:
        explode = ((0,) * (len(finalList)))
        
    patches, texts, autotexts = ax.pie(finalList, explode=explode, colors=colors, labels=labelsQ, autopct='%1.1f%%',
           shadow=False, startangle=90, wedgeprops={ 'linewidth' : 1, 'edgecolor' : "white" },  pctdistance=pctdistance,
            )

#    ax.set_title(field, y=1.08)

    # PIE TITLE
    q = str(sum([int(q) for q in quantities]))

    if pieTitle:
        defaultTitle = givenTitle
        if field != None:
            defaultTitle = field
        if givenTitle == None:
            defaultTitle = columnname
#            defaultTitle = columnname.split()[0]
        titleWithQ = defaultTitle + ' (total: ' + q + ')'
        fig.suptitle(titleWithQ, fontsize=20, y=0.95, **csfont)
        
    for t in texts:
        t.set_size('xx-large')
    for t in autotexts:
        t.set_size('x-large')
        t.set_color('white')
   
   
    # LEGEND
    legend = plt.legend(labels=labels, loc=10, bbox_to_anchor=(0.5, -0.08), ncol=2)
    if legendloc == 'Side':
        legend = plt.legend(labels=labels, loc=10, bbox_to_anchor=(1.3, 0.65), ncol=1)
    legend.get_frame().set_edgecolor('#131339')
    legend.get_frame().set_alpha(0.9)
    plt.rcParams['legend.fancybox'] = True
    plt.rcParams['legend.shadow'] = True
    plt.rcParams['legend.borderpad'] = 1.5
    plt.rcParams['legend.handlelength'] = 2

#    plt.show()  # Actually, don't show, just save to foo.png
#    
#    # SAVE
    name = columnname
    if field !=None:
        name = field
    regex = re.compile('[^a-zA-Z0-9]')
    altered = regex.sub('_', name)
    savename = altered + '_Pie.png'
    current_dir(currdir)
#    change_dir()
    fig.savefig(savename, dpi=199, bbox_inches='tight')

#-------------------------------------------------------------------------------------#

# TKINTER

# ======== Select a directory:

#import Tkinter, tkFileDialog
#
#from tkFileDialog import askopenfilename
#from Tkinter import *
#import ttk
#
#content = ''
#file_path = ''
#ftypes = [('Excel File', '*.xlsx'), ('All files', '*')]
#
##~~~~ FUNCTIONS~~~~
#
#def open_file():
#    global content
#    global file_path
#    get_plot_style()
##    filename = askopenfilename()
#    filename = tkFileDialog.askopenfilename(filetypes=ftypes)
#    if filename: 
#        try: 
##            self.settings["template"].set(filename)
#            file_path = filename
##            print file_path
#        except: 
#            messagebox.showerror("Open Source File", "Failed to read file \n'%s'"%filename)
#            return
#        
##    infile = open(filename, 'r')
##    content = infile.read()
##    file_path = os.path.dirname(filename)
#    entry.delete(0, END)
#    entry.insert(0, file_path)
##    return filename
#
#def process_file(filename):
#    test(filename)
#    
#def exit_program():
#    pass
#
#def test():    
#    directory = file_path.split('/')
#    selected = directory.pop() 
##    cwd = '/'.join(directory)
#    openpyxl.load_workbook(selected)
#    curr = os.getcwd()

#  #~~~~~~~~~~~~~~~~~~~
#
# #~~~~~~ GUI ~~~~~~~~
#
#root = Tk()
#root.title('Concession Charts Tool')
#root.geometry("598x280")
#
#mf = Frame(root, pady=10)
##mf.grid(row=2, column=2)
#mf.pack()
#
#s = ttk.Style()
#s.theme_use("clam")
#print s.theme_use()
#
## row 1 ------------------------------------------
#f1 = Frame(mf, width=600, height=250)
#f1.pack()
#Label(f1,text="Select Your File (Only Excel files):").grid(row=0, column=0, sticky='e')
#Entry(f1, width=50, textvariable=file_path).grid(row=0,column=1,padx=2,pady=2,sticky='we',columnspan=25)
#Button(f1, text="Browse", command=open_file).grid(row=0, column=27, sticky='ew', padx=8, pady=4)
## Separator ------------------------------------------
#ttk.Separator(root).place(y=50, relwidth=1)
## row 2 ------------------------------------------
#f2 = Frame(mf, width=400, height=300)
#f2.pack(side=LEFT, pady=15)
## Combobox
#choices = ['create barchart','create pie chart','option3','option4']
#box_value = StringVar(root)
#box_value.set(choices[0])
#Label(f2,text="Please make a selection:").grid(row=1, column=0, sticky='ew')
#box = OptionMenu(f2, box_value, *choices).grid(row=1, column=1)
## row 3 ------------------------------------------
#
## row 4 ------------------------------------------
#
## row 5 ------------------------------------------
#
### row 6 ------------------------------------------
##f3 = Frame(mf, width=400, height=300)
##f3.pack()
##f4 = Frame(mf, width=600, height=250)
##f4.pack()
##f5 = Frame(mf, width=600, height=250)
##f5.pack()
##f8 = Frame(mf, width=400, height=250).pack()
#
##t = Toplevel(sup, borderwidth=2, relief = relief)
##Label(t, text=relief, width=10).pack(side=LEFT)
#
##listbox = Listbox(f8)
##listbox.pack()
##listbox.insert(END, 'a list entry')
##for item in ["one", "two", "three", "four"]:
##    listbox.insert(END, item)
#
#
##file_path = StringVar
##var = tkSimpleDialog.askstring("Name prompt", "enter your name")
#
##dlg = tkFileDialog.Open(self, filetypes = ftypes)
##fl = dlg.show()
#
#
#
#
##Label(f2,text="Column X").grid(row=0, column=0, sticky='e')
##Entry(f2, width=30, textvariable=file_path).grid(row=0,column=1,padx=2,pady=2,sticky='e',columnspan=25)
##
##Label(f3,text="Column Y").grid(row=0, column=0, sticky='e')
##Entry(f3, width=30, textvariable=file_path).grid(row=0,column=1,padx=2,pady=2,sticky='e',columnspan=25)
##
#
##Button(f4, text="Process Now", width=32, command=lambda: test()).grid(sticky='ew', padx=10, pady=10)
##Button(f5, text="exit", width=25, command=root.destroy).grid(row=0, column=0, padx=4, pady=4)
#
#
#
##button = Button(root, text="check value slected")
##button.pack(side='left', padx=20, pady=10)
#

#root.mainloop()


    
 


